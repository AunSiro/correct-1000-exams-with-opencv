
from openpyxl import Workbook
from openpyxl import load_workbook
import os



num = 0
something_found = True
while something_found:
    num += 1
    root = 'correcciones/resultado' + str(num+1).zfill(3) + '.xlsx'
    something_found = os.path.exists(root)

print('corrigiendo ', num, ' exámenes encontrados')
#Cargar clave
wb_clave = load_workbook(filename='clave.xlsx', read_only=True)

#Seleccionar la worksheet activa
ws_clave = wb_clave.active

#Convertir a lista de listas    
datos_clave = []
for row in ws_clave.rows:
    fila = []
    for cell in row:
        fila.append(cell.value)
    datos_clave.append(fila)
    
    
#Crear workbook nuevo
wb_results = Workbook()

#Seleccionar la worksheet activa
ws_results = wb_results.active    
ws_results.append(['examen', 'aciertos', 'blancos', 'fallos'])  

#Corregir cada examen  
for ii in range(num):
    root =  'correcciones/resultado' + str(ii+1).zfill(3) + '.xlsx'
    aciertos = 0
    fallos = 0
    blancos = 0

    #Crear workbook nuevo
    wb = load_workbook(filename = root, read_only=True)
    

    #Seleccionar la worksheet activa
    ws = wb.active
    
    
    #Comparar cada pregunta con la solucion correcta
    datos = []
    for row in ws.rows:
        fila = []
        for cell in row:
            fila.append(cell.value)
        datos.append(fila)
    
    for fila in range(len(datos)):
        prop = datos[fila][1]
        
        if prop == datos_clave[fila][1]:
            aciertos += 1
        elif prop == None:
            blancos += 1
        else:
            fallos += 1
    #Guardar los datos        
    ws_results.append([ii+1, aciertos, blancos, fallos])    
    del(ws)
    del(wb)
wb_results.save('resultados.xlsx')
print('\n\nCorrección completada.\n\n')

