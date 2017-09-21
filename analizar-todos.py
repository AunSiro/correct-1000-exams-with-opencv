from analisis import evaluar
from openpyxl import Workbook
from openpyxl import load_workbook
import os

must_repeat = True
something_found = False
num_found = 0

#interfaz
while must_repeat:
    print('introduzca nombre del archivo genérico')
    nombre = input()
    try_name_1 = 'examenes/' + nombre + '001.jpg'
    try_name_2 = nombre + '001.jpg'
    
    if os.path.exists(try_name_1):
        something_found = True
        pre_name = 'examenes/'
    elif os.path.exists(try_name_2):        
        something_found = True
        pre_name = ''
    else:
        print('Ningun archivo encontrado llamado ', try_name_2)
        continue
    
    while something_found:
        num_found += 1
        root = pre_name + nombre + str(num_found+1).zfill(3) + '.jpg'
        something_found = os.path.exists(root)
    print('Encontrados ', num_found, ' archivos, ¿Continuar?')
    resp = input()
    if len(resp)>0:
        resp = resp[0].lower()
    else:
        #Si el usuario pulsa intro sin escribir nada,
        #se considera como autorizado a continuar
        resp = 's'
    if resp in 'sy':
        print('continuar analizando')
        must_repeat = False
    else:
        print('cancelado por usuario, probar nombre nuevo')
        
    print()
        
 

#Crear workbook nuevo
wb_fallos = Workbook()

#Seleccionar la worksheet activa
ws_fallos = wb_fallos.active

#Escribir cabecera
ws_fallos.append(['examen', 'pregunta','propuesta', 'error'])

#Si no hay carpeta, se crea 
if not os.path.isdir('correcciones'):
    os.makedirs('correcciones')

if not os.path.isdir('examenes_coloreados'):
    os.makedirs('examenes_coloreados')

#Analizar cada examen        
for ii in range(num_found):
    root = pre_name + nombre + str(ii+1).zfill(3) + '.jpg'
    print('Evaluando archivo: ', root)
    saveroot = 'correcciones/resultado' + str(ii+1).zfill(3) + '.xlsx'

    result = evaluar(root)
    
    #Crear workbook nuevo
    wb = Workbook()

    #Seleccionar la worksheet activa
    ws = wb.active
    
    #Escribir resultados del examen
    for line in result: 
        #Cada linea del examen se escribe en el wb del examen
        ws.append(line)
        if len(line[2]) > 1:
            #Los fallos de cada examen se añaden al wb de fallos
            ws_fallos.append([ii+1, line[0], line[1], line[2]])
    wb.save(saveroot)
    del(ws)
    del(wb)

wb_fallos.save('fallos.xlsx')
del(ws_fallos)
del(wb_fallos)
print('\n\nAnálisis completado.\n\n')